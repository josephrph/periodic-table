#!/usr/bin/env python3
"""
preflight.py — run this BEFORE every deploy. Exit 0 = safe to push.

Tier 2 of the self-maintenance plan (see UX-62 in Project_Backlog.xlsx). Tier 1 (V2FACTS)
stops counts from drifting by deriving them from the data. This script catches the classes
of problem that derivation cannot: stale fallbacks, malformed entities, broken data
references, banned terminology, and violations of the standing product rules in CLAUDE.md.

Every check here exists because the corresponding bug ACTUALLY SHIPPED at least once:

  fallback literals    "51 to 65" survived the cannabicitran merge (65 -> 64)
  double-escaped HTML  the water example read H&amp;sub2;O and could never render
  unknown entities     &sub2; is not a real entity even unescaped
  FAQ question drift   Essential Tremor was listed in one FAQ copy and not the other
  banned terminology   "PhytoTable" and "Clinically Structured" both had to be purged
  trademark lockups    the (TM) mark was missing from 13 brand lockups
  duplicate backlog ID two rows were filed as DRUG-A3 and two as MOB-07

    python3 preflight.py              # offline checks only (fast, no network)
    python3 preflight.py --online     # additionally verify every PMID against NCBI

Add a check whenever a content or data bug gets past review. That is the whole point:
this file is the project's memory of its own mistakes.
"""
import html.entities
import json
import os
import re
import subprocess
import sys
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, 'index.html')
JSC = ('/System/Library/Frameworks/JavaScriptCore.framework/Versions/Current/'
       'Helpers/jsc')

FAILURES = []
NOTES = []


# ── projections of the source ──────────────────────────────────────────────────
# Several checks must look only at what a reader can actually SEE. Without this, the
# first run flagged 14 entity "bugs" that were really search-index attribute values, a
# banned typo that only appears in a comment documenting its own removal, and a lockup
# inside a comment quoting CLAUDE.md. A checker that cries wolf gets switched off.
def strip_comments(src):
    """Blank out HTML, CSS and JS comments, preserving offsets so context stays readable."""
    def blank(m):
        return re.sub(r'[^\n]', ' ', m.group(0))
    src = re.sub(r'<!--.*?-->', blank, src, flags=re.S)
    src = re.sub(r'/\*.*?\*/', blank, src, flags=re.S)
    # line comments: only when // starts a line (avoids gutting "https://" and regexes)
    src = re.sub(r'(?m)^(\s*)//[^\n]*', lambda m: m.group(1) + ' ' * (len(m.group(0)) - len(m.group(1))), src)
    return src


def strip_attributes(src):
    """Blank out double-quoted attribute values. Attributes are never rendered as text."""
    return re.sub(r'="[^"]*"', lambda m: '="' + ' ' * (len(m.group(0)) - 3) + '"', src)


def group_keys(raw):
    """A condition's group is usually a string, but a cross-listed one is an array.

    prostate-cancer is ['cancer-sub','mens']. JS gets this right for free because
    String(array) joins on commas; Python's str(list) does not, which made this script
    report 10 cancer sub-types where the app correctly shows 11.
    """
    if isinstance(raw, (list, tuple)):
        parts = raw
    else:
        parts = str(raw or '').split(',')
    return [str(p).strip() for p in parts if str(p).strip()]


def fail(check, msg):
    FAILURES.append((check, msg))


def note(msg):
    NOTES.append(msg)


# ── extracting the live data ───────────────────────────────────────────────────
def js_literal(src, marker, opener):
    """Brace/bracket-match a JS literal out of index.html, skipping strings and comments.

    The naive regex version of this silently dropped 32 of 65 molecules once, and a later
    version desynchronised on an apostrophe inside a // comment. Hence the explicit scanner.
    """
    i = src.index(marker)
    j = src.index(opener, i)
    closer = {'{': '}', '[': ']'}[opener]
    depth, k, in_str, quote, esc = 0, j, False, '', False
    while k < len(src):
        c = src[k]
        nxt = src[k + 1] if k + 1 < len(src) else ''
        if in_str:
            if esc:
                esc = False
            elif c == '\\':
                esc = True
            elif c == quote:
                in_str = False
        elif c == '/' and nxt == '/':
            k = src.find('\n', k)
            if k < 0:
                break
            continue
        elif c == '/' and nxt == '*':
            k = src.find('*/', k)
            if k < 0:
                break
            k += 2
            continue
        elif c in '\'"`':
            in_str, quote = True, c
        elif c == opener:
            depth += 1
        elif c == closer:
            depth -= 1
            if depth == 0:
                break
        k += 1
    return src[j:k + 1]


def load_data(src):
    """Evaluate the app's own data objects in JavaScriptCore and return them as Python."""
    parts = {
        'M': js_literal(src, 'const M = {', '{'),
        'CONDITIONS': js_literal(src, 'CONDITIONS = [', '['),
        # DRUG-09: the drug tables are evaluated too. A regex-based audit of these reported 11
        # unreachable drug-drug pairs that did not exist — DI_DATA is written in TWO styles,
        # single-quoted JS and double-quoted JSON, and the pattern only saw the first. Same class
        # of mistake as the citation-array regex that produced a false "42 conditions uncited".
        'DI_DATA': js_literal(src, 'var DI_DATA = [', '['),
        'DDI_DATA': js_literal(src, 'var DDI_DATA = [', '['),
        # DRUG-24: needed by check_hasrisk_invariant
        'ADVERSE_FINDINGS': js_literal(src, 'const ADVERSE_FINDINGS = {', '{'),
    }
    prog = ['var out = {};']
    for name, body in parts.items():
        prog.append('out.%s = %s;' % (name, body))
    prog.append('print(JSON.stringify(out));')
    tmp = os.path.join(HERE, '_preflight_tmp.js')
    with open(tmp, 'w', encoding='utf-8') as fh:
        fh.write('\n'.join(prog))
    try:
        res = subprocess.run([JSC, tmp], capture_output=True, text=True, timeout=90)
    finally:
        os.remove(tmp)
    if res.returncode != 0 or not res.stdout.strip().startswith('{'):
        raise SystemExit('could not evaluate app data:\n' + (res.stderr or res.stdout)[:600])
    return json.loads(res.stdout)


# ── checks ─────────────────────────────────────────────────────────────────────
def check_scripts_parse(src):
    """Every inline <script> must still compile."""
    blocks = re.findall(r'<script(?![^>]*src=)[^>]*>(.*?)</script>', src, re.S)
    for i, body in enumerate(blocks):
        if not body.strip():
            continue
        tmp = os.path.join(HERE, '_preflight_blk%d.js' % i)
        with open(tmp, 'w', encoding='utf-8') as fh:
            fh.write(body)
        try:
            res = subprocess.run(
                [JSC, '-e',
                 'var s=readFile("%s");try{new Function(s);print("OK");}'
                 'catch(e){print("FAIL "+e);}' % tmp],
                capture_output=True, text=True, timeout=90)
        finally:
            os.remove(tmp)
        out = (res.stdout or '').strip()
        if not out.startswith('OK'):
            fail('scripts parse', 'script block %d: %s' % (i, out or res.stderr.strip()[:200]))


def check_v2fact_keys(src, data):
    """Each data-v2fact key must be a real V2FACTS function, and its fallback must be current.

    A typo'd key fails silently — the span just keeps its literal forever, which is exactly
    the hand-maintained state Tier 1 was meant to remove. A stale literal is worse: it is
    the pre-UX-61 bug wearing a placeholder's clothes.
    """
    mols = data['M']
    conds = [c for c in data['CONDITIONS'] if c.get('id') != 'cancer-adverse']
    n_acid = sum(1 for k in mols if mols[k].get('category') == 'acid')
    groups = {}
    for c in conds:
        for g in group_keys(c.get('group')):
            groups.setdefault(g, []).append(c.get('label'))
    expected = {
        'molecules': len(mols),
        'acids': n_acid,
        'nonAcid': len(mols) - n_acid,
        'conditions': len(conds),
        'cancerSubtypes': len(groups.get('cancer-sub', [])),
        'v1Entries': 51,
        'addedSinceV1': len(mols) - 51,
    }
    # UX-74: receptor tallies. These sat in FAQ prose by hand and had drifted (CB1 read 16
    # against 17 in the data, CB2 read 23 against 22), so they are derived and asserted now.
    for rc in ('CB1', 'CB2', 'TRP', 'PPAR', '5HT'):
        expected['rec' + rc] = sum(
            1 for k in mols if rc in (mols[k].get('receptors') or []))
    found = re.findall(r'<span data-v2fact="([^"]+)"[^>]*>([^<]*)</span>', src)
    if not found:
        fail('V2FACTS wiring', 'no data-v2fact spans found at all')
    for key, literal in found:
        if key not in expected:
            fail('V2FACTS wiring',
                 'data-v2fact="%s" has no matching V2FACTS function' % key)
            continue
        want = str(expected[key])
        if literal.strip() != want:
            fail('V2FACTS fallback',
                 'data-v2fact="%s" fallback reads "%s" but the data says %s'
                 % (key, literal.strip(), want))
    # the generated containers must declare a style the module understands
    for style in re.findall(r'data-v2fact="conditionGroups"[^>]*data-v2fact-style="([^"]*)"', src):
        if style not in ('short', 'long'):
            fail('V2FACTS wiring', 'conditionGroups style "%s" is not short|long' % style)

    # UX-120: the conditionGroups fallbacks were NOT checked. check_v2fact_keys validated every
    # <span> fallback but these are <div>s, and only their style attribute was asserted — so while
    # every scalar stayed correct, the two group listings silently drifted four conditions behind
    # the data (Pain read 7 against 8, Neurological 12 against 14, Cancer & Immune 5 against 6;
    # Sickle Cell Disease, Spinal Cord Injury, Traumatic Brain Injury and HIV/AIDS were all
    # missing). Users never saw it because V2FACTS.apply() overwrites the container at load, but a
    # stale literal in the build is exactly what Tier 1 exists to prevent. Membership and counts
    # are asserted here, not markup, so wording and punctuation stay free to change.
    import html as _html
    label_of = {}
    for c in conds:
        for g in group_keys(c.get('group')):
            label_of.setdefault(g, []).append(c.get('label'))
    for m in re.finditer(r'<div data-v2fact="conditionGroups"[^>]*data-v2fact-style="(short|long)"[^>]*>([\s\S]*?)</div>', src):
        style, body = m.group(1), m.group(2)
        paras = re.findall(r'<u>([^<]*?)\s*\((\d+)[^)]*\):</u>([^<]*)', body)
        if len(paras) != len(label_of):
            fail('V2FACTS fallback',
                 'conditionGroups[%s] lists %d groups but the data has %d'
                 % (style, len(paras), len(label_of)))
            continue
        for gname, gcount, members in paras:
            gname = _html.unescape(gname).strip()
            listed = [x.strip() for x in _html.unescape(members).strip().rstrip('.').split(',') if x.strip()]
            if int(gcount) != len(listed):
                fail('V2FACTS fallback',
                     'conditionGroups[%s] "%s" says (%s) but lists %d conditions'
                     % (style, gname, gcount, len(listed)))
            match = [g for g in label_of if len(label_of[g]) == int(gcount)
                     and sorted(x.replace(' / ', '/') for x in label_of[g])
                         == sorted(x.replace(' / ', '/') for x in listed)]
            if not match:
                want = [(g, len(label_of[g])) for g in label_of]
                fail('V2FACTS fallback',
                     'conditionGroups[%s] "%s (%s)" does not match any group in the data; '
                     'data groups are %s' % (style, gname, gcount, want))
    note('V2FACTS: conditionGroups fallbacks match the data (%d groups)' % len(label_of))
    note('V2FACTS: %d spans checked against live data %s' % (len(found), expected))


def check_demo_guided_parity(src, data):
    """UX-123: Demo Mode must not keep private copies of guided data.

    Demo Mode and Guided Match are two flows over one dataset. Where they duplicated a table
    instead of sharing it, they drifted — GUIDED_MOL was copied wholesale into DJ_GUIDED_MOL
    because a closure boundary put the original out of reach, and the science-alias map was
    written out five separate times. This asserts the de-duplication holds and that every
    condition the demo offers is a real one.
    """
    body = strip_comments(src)

    # 1. no second copy of the curated guided-molecule profiles
    if 'DJ_GUIDED_MOL' in body:
        fail('demo parity',
             'DJ_GUIDED_MOL is back — the demo must read window.V4Guided.guidedMol, not its own copy')

    # 2. exactly one literal of the science-alias map
    n_alias = len(re.findall(r"'appetite-stim'\s*:\s*'appetite'", body))
    if n_alias != 1:
        fail('demo parity',
             'the appetite/nausea science-alias map appears %d times; it must be declared once'
             % n_alias)

    # 3. every id the demo offers must resolve — a real condition, or one of the curated
    #    guided-only profiles that the science layer aliases back to a real condition
    m = re.search(r'var DJ_CONDS\s*=\s*\[(.*?)\];', body, re.S)
    if not m:
        fail('demo parity', 'DJ_CONDS not found')
        return
    dj_ids = re.findall(r"id:'([^']+)'", m.group(1))
    real = {c.get('id') for c in data['CONDITIONS']}
    curated = set(re.findall(r"'([a-z-]+)'\s*:\s*\{[A-Z]", 
                             re.search(r'var GUIDED_MOL=\{(.*?)\};', body, re.S).group(1))) \
              if re.search(r'var GUIDED_MOL=\{(.*?)\};', body, re.S) else set()
    orphans = [i for i in dj_ids if i not in real and i not in curated]
    if orphans:
        fail('demo parity',
             'Demo Mode offers condition id(s) that resolve to nothing: %s' % ', '.join(orphans))
    note('demo parity: %d demo conditions all resolve (%d real, %d curated profiles); '
         'no duplicated guided tables'
         % (len(dj_ids), len([i for i in dj_ids if i in real]),
            len([i for i in dj_ids if i in curated])))


def check_entities(src):
    """No double-escaped entities, and no entity names the browser will not know.

    Scoped to rendered text: attribute values legitimately contain escaped entity text
    (the tile search index stores "&delta;9-thc" so a search for "delta" still matches),
    and comments are not rendered at all.
    """
    src = strip_attributes(strip_comments(src))
    for m in re.finditer(r'&amp;([a-zA-Z][a-zA-Z0-9]{1,12});', src):
        ctx = re.sub(r'\s+', ' ', src[max(0, m.start() - 45):m.end() + 25])
        fail('HTML entities',
             'double-escaped "&amp;%s;" — renders literally: …%s…' % (m.group(1), ctx))
    known = set(html.entities.entitydefs)
    for m in re.finditer(r'&([a-zA-Z][a-zA-Z0-9]{1,12});', src):
        name = m.group(1)
        if name not in known:
            ctx = re.sub(r'\s+', ' ', src[max(0, m.start() - 45):m.end() + 25])
            fail('HTML entities', 'unknown entity "&%s;" …%s…' % (name, ctx))


def check_banned_terms(src):
    """Terminology the owner has explicitly retired. See CLAUDE.md standing rules.

    Comments are excluded: a note explaining that a typo was fixed is not the typo.
    """
    src = strip_comments(src)
    banned = {
        'PhytoTable': 'retired product name (must not appear in any user-facing text)',
        'Clinically Structured': 'removed from the tagline (UX-38)',
        'Perioidic': 'typo that shipped once',
        'Cannbicitran': 'typo that shipped once',
    }
    for term, why in banned.items():
        # the staff-only export filename is an accepted internal exception
        hits = [m.start() for m in re.finditer(re.escape(term), src, re.I)
                if 'phytotable_feedback' not in src[max(0, m.start() - 30):m.start() + 40].lower()]
        if hits:
            fail('banned terminology', '"%s" appears %d time(s) — %s' % (term, len(hits), why))


def check_brand_lockups(src):
    """Standing rule: every brand lockup carries the trademark mark (UX-54).

    Comments excluded — CLAUDE.md's own wording is quoted in a CSS comment.
    """
    src = strip_comments(src)
    pat = re.compile(r'(?:The Periodic Table of (?:<em>)?Cannabis(?:</em>)? Plant Molecules)'
                     r'(&trade;|&#8482;|™)?')
    missing = 0
    for m in pat.finditer(src):
        if not m.group(1):
            ctx = re.sub(r'\s+', ' ', src[max(0, m.start() - 60):m.end() + 20])
            fail('trademark lockups', 'lockup without the mark: …%s…' % ctx)
            missing += 1
    if not missing:
        note('trademark: every "Periodic Table of Cannabis Plant Molecules" lockup carries the mark')


def check_standing_rules(src):
    """The non-negotiables from CLAUDE.md that a refactor could quietly break."""
    required = {
        'id="guidedBtn"': 'persistent Guided Match header button',
        'id="mtbGuided"': 'mobile Guided tab',
        'Mechanism Based &#183; Evidence Informed &#183; Easy to Explore':
            'header tagline (UX-38 wording)',
        'id="inactivityWarn"': 'inactivity warning element',
        'id="inactivityCountdown"': 'inactivity countdown element',
        'is-mobile-view': 'mobile view class',
    }
    for needle, what in required.items():
        probe = needle.replace('&#183;', '·')
        if needle not in src and probe not in src:
            fail('standing rules', 'missing %s (%s)' % (needle, what))


def check_data_integrity(src, data):
    """The molecule/condition graph must not reference anything that does not exist."""
    mols, conds = data['M'], data['CONDITIONS']
    tiles = re.findall(r'<div class="mol-tile [^"]*"[^>]*?data-id="([^"]+)"', src)
    if len(tiles) != len(mols):
        fail('data integrity',
             '%d .mol-tile divs but %d entries in M' % (len(tiles), len(mols)))
    for t in tiles:
        if t not in mols:
            fail('data integrity', 'tile data-id="%s" is not in M' % t)
    # one tile per grid cell
    cells = re.findall(r'grid-area:\s*(\d+)\s*/\s*(\d+)', src)
    dupes = {c for c in cells if cells.count(c) > 1}
    if dupes:
        fail('data integrity', 'two tiles share grid cell(s): %s' % sorted(dupes))
    # duplicate symbols read as a data-entry slip (a real one shipped: cannabicitran)
    syms = [mols[k].get('symbol') for k in mols]
    dup_syms = sorted({s for s in syms if syms.count(s) > 1})
    if dup_syms:
        fail('data integrity', 'duplicate molecule symbols: %s' % dup_syms)
    # every condition points at molecules that exist, and has at least one
    for c in conds:
        cid = c.get('id')
        mset = c.get('molecules') or {}
        if cid != 'cancer-adverse' and not mset:
            fail('data integrity', 'condition "%s" has no molecules' % cid)
        for mid in mset:
            if mid not in mols:
                fail('data integrity',
                     'condition "%s" references unknown molecule "%s"' % (cid, mid))


def check_faq_parity(src):
    """The two FAQ copies must answer the same set of questions.

    They are allowed to differ in depth — the overlay is a condensed tier — but a question
    present in one and absent from the other is how Essential Tremor went missing.
    """
    i = src.index('id="wlc-faqs"')
    i = src.rindex('<div', 0, i)
    depth, k = 0, i
    while k < len(src):
        if src.startswith('<div', k):
            depth += 1
        elif src.startswith('</div>', k):
            depth -= 1
            if depth == 0:
                k += 6
                break
        k += 1
    overlay = src[i:k]
    b0 = src.index('<section id="faq-section">')
    main = src[b0:src.index('</section>', b0)]

    def questions(blk):
        out = []
        for m in re.finditer(r'<span class="faq-q"[^>]*>(.*?)</span>', blk, re.S):
            q = re.sub(r'\s+', ' ', re.sub(r'<[^>]+>', '', m.group(1))).strip().lower()
            out.append(re.sub(r'[^a-z0-9 ]', '', q))
        return out

    qa, qb = questions(overlay), questions(main)
    # compare on a loose key so wording differences ("…of this table") do not trip it
    def key(q):
        drop = {'the', 'a', 'an', 'of', 'this', 'in', 'and', 'is', 'are', 'does', 'do',
                'what', 'how', 'were', 'v2', 'table', 'chart', 'my', 'me', 'i', 'it'}
        return ' '.join(sorted(w for w in q.split() if w not in drop))
    ka, kb = {key(q) for q in qa}, {key(q) for q in qb}
    for k2 in sorted(ka - kb):
        if k2:
            note('FAQ: only in the How-to-Use overlay — "%s"' % k2)
    for k2 in sorted(kb - ka):
        if k2:
            note('FAQ: only in the main FAQ — "%s"' % k2)
    note('FAQ parity: overlay %d questions, main %d, %d shared'
         % (len(qa), len(qb), len(ka & kb)))


def check_pmids(src, online=False):
    """PMIDs must be plausible, and optionally must actually resolve at NCBI."""
    # BOTH quote styles: the curated citation objects use pmid:'...' while the JSON-shaped
    # DI_DATA records use "pmid":"...". Matching only the first undercounted by 15.
    # AUDIT-09: this used to be r'[\'"](\d{5,8})[\'"]', which requires the closing quote to sit
    # immediately after the digits. That silently skipped EVERY multi-PMID field — 'pmid' values
    # holding several ids separated by commas, which is how the drug tables cite more than one
    # paper — and it also let 'N/A - theoretical' through without comment. Capture the whole field,
    # then split it the way diPmidList() does in the app, so preflight verifies exactly the set of
    # ids the page will turn into links.
    raw_fields = re.findall(r'"?pmid"?\s*:\s*[\'"]([^\'"]*)[\'"]', src)
    pmids, junk = set(), set()
    for field in raw_fields:
        for tok in re.split(r'[,;]', field):
            tok = tok.strip()
            if not tok:
                continue
            if re.fullmatch(r'\d{5,8}', tok):
                pmids.add(tok)
            else:
                junk.add(field.strip())
    pmids = sorted(pmids)
    if junk:
        note('citations: %d pmid field(s) carry non-PMID text and render no link (by design): %s'
             % (len(junk), sorted(junk)[:2]))
    bad = [p for p in pmids if not (5 <= len(p) <= 8)]
    if bad:
        fail('citations', 'implausible PMIDs: %s' % bad[:12])
    note('citations: %d distinct PMIDs referenced' % len(pmids))
    if not online:
        return
    missing = []
    for i in range(0, len(pmids), 180):
        batch = pmids[i:i + 180]
        url = ('https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi'
               '?db=pubmed&retmode=json&id=' + ','.join(batch))
        try:
            with urllib.request.urlopen(url, timeout=60) as r:
                res = json.load(r).get('result', {})
        except Exception as exc:                     # network trouble is not a content bug
            note('citations: NCBI lookup skipped (%s)' % exc)
            return
        for p in batch:
            rec = res.get(p)
            if not rec or rec.get('error') or not rec.get('title'):
                missing.append(p)
    if missing:
        fail('citations', '%d PMID(s) did not resolve at NCBI: %s'
             % (len(missing), missing[:12]))
    else:
        note('citations: all %d PMIDs resolved at NCBI' % len(pmids))


def check_terminology(src):
    """Tier 3: category labels must come from V2TERMS, and the tagline must be canonical.

    Five separate label maps had drifted so that the 'acid' category answered to four
    different names in four places. Any new literal map is how that happens again.
    """
    src_nc = strip_comments(src)
    # a label map is a literal object keyed by the five category ids
    for m in re.finditer(r'(?:const|var|let)\s+(\w*[Cc]at[A-Za-z_]*)\s*=\s*\{([^}]*)\}', src_nc):
        body = m.group(2)
        if 'phytocannabinoid' in body and ('terpene' in body and 'flavonoid' in body):
            # a LABEL map holds words; catColor holds hex values — only flag the former
            vals = re.findall(r":\s*'([^']*)'", body)
            if vals and not any(v.startswith('#') for v in vals):
                fail('terminology',
                     'literal category-label map "%s" — derive it from V2TERMS.map() instead'
                     % m.group(1))
    # the tagline is fixed by CLAUDE.md standing rule #1
    canon_words = ['Mechanism Based', 'Evidence Informed', 'Easy to Explore']
    for m in re.finditer(r'Mechanism[ \-]Based[^<"\']{0,70}', src_nc):
        text = m.group(0)
        plain = text.replace('&middot;', '·')
        if not all(w in plain for w in canon_words):
            fail('terminology',
                 'tagline variant "%s" — CLAUDE.md fixes the wording as "%s"'
                 % (plain.strip(), 'Mechanism Based · Evidence Informed · Easy to Explore'))
    if 'window.V2TERMS' not in src:
        fail('terminology', 'V2TERMS is missing — category labels have no single source')
    else:
        note('terminology: category labels sourced from V2TERMS; tagline wording canonical')


def check_recommendation_wording(src):
    """UX-66: the recommendation heading and its attribution must stay uniform.

    The old wording ("What the research associates", "The molecules V2 associates with") sat in
    four places — two screens and two FAQ copies that quote the screen title. Renaming only the
    screens would leave the FAQ describing a title that no longer exists, which is the same drift
    class as the counts before V2FACTS.
    """
    src_nc = strip_comments(src)
    retired = {
        'What the research associates': 'old recommendation screen heading',
        'V2 associates': 'attributed the association to V2 rather than the literature',
        'research associates': 'superseded by "research links to"',
    }
    for phrase, why in retired.items():
        hits = len(re.findall(re.escape(phrase), src_nc, re.I))
        if hits:
            fail('recommendation wording',
                 '"%s" appears %d time(s) — %s' % (phrase, hits, why))
    heading = 'Research-Linked Molecules for Your Health Topic'
    n = len(re.findall(re.escape(heading), src_nc))
    # 2 rendered screens (standard + Demo Mode) and 2 FAQ copies quoting the title
    if n != 4:
        fail('recommendation wording',
             'heading "%s" appears %d time(s), expected 4 (2 screens + 2 FAQ quotes)' % (heading, n))
    else:
        note('recommendation wording: heading present in all 4 expected places, no retired phrasings')
    # any topic flagged evidenceNegative must actually carry adverse-findings copy, or the
    # softened wording would be the only thing telling the reader the trials were null
    for m in re.finditer(r"\{id:'([a-z0-9\-]+)',label:'([^']*)'[^}]*?evidenceNegative:true", src):
        if m.group(2) not in src:
            fail('recommendation wording',
                 'condition "%s" is flagged evidenceNegative but has no findings copy' % m.group(2))


def check_greek_notation(src):
    """UX-70: Greek prefixes are written as symbols, and BOTH notations stay searchable.

    Three separate traps make this worth asserting rather than trusting:

    1. Quoted literature must keep the publisher's spelling. A citation title reading
       "Delta-9-tetrahydrocannabinol enhances breast cancer growth" is what the journal
       printed; rewriting it to Δ9 falsifies the reference. So this check deliberately
       ignores `title:` fields and only polices V2's OWN name fields.
    2. Search keys must survive the change. `data-name` and PUBMED_TERMS are haystacks, not
       labels — indexing only "&Delta;9-..." would break the very common query "delta-9"
       and send PubMed a term that returns nothing.
    3. Adrenergic receptor subtypes and drug classes are NOT cannabis molecules.
       "beta-blockers", "Beta-Adrenergic" and "Beta-lactamase" are the standard clinical
       spellings and a blanket replace would corrupt them.
    """
    # 1. no molecule name field may spell a prefix out
    spelled = []
    for m in re.finditer(r"fullName:'([^']*)'", src):
        if re.search(r'\b(Delta|Alpha|Beta|Gamma)-', m.group(1)):
            spelled.append(m.group(1))
    if spelled:
        fail('greek', 'fullName still spells out a Greek prefix: %s' % ', '.join(spelled[:4]))

    # 2. both notations must remain searchable
    if 'pubmedName(mol.fullName)' not in src:
        fail('greek', 'tile search keys no longer carry the spelled-out alias — '
                      '"delta-9" would stop matching')
    # bound the map at its own closing brace — a lazy ".*?\n  };" runs straight past it
    # into unrelated code and then reports Greek found in a comment.
    i = src.find('const PUBMED_TERMS = {')
    if i == -1:
        fail('greek', 'PUBMED_TERMS map is missing; PubMed links lose their curated terms')
    else:
        body = src[i:src.index('};', i)]
        if re.search(r'[\u0394\u03b1\u03b2]', body):
            fail('greek', 'PUBMED_TERMS contains a Greek symbol; PubMed indexes the '
                          'spelled-out form')

    # 3. clinical terms that must NOT be converted
    for term in ('Beta-Blockers', 'Beta-Adrenergic', 'Beta-lactamase'):
        if term not in src:
            fail('greek', 'clinical term "%s" was wrongly converted to a Greek symbol' % term)

    n_titles = len([t for t in re.findall(r"title:'((?:[^'\\]|\\.)*)'", src)
                    if re.search(r'\b(Delta|Alpha|Beta)-', t)])
    note('greek notation: name fields use symbols; %d citation titles correctly left as '
         'published; spelled-out search aliases intact' % n_titles)


def check_pinch_zoom(src):
    """UX-71: no rule may withhold pinch-zoom, and the viewport may not forbid scaling.

    Accessibility floor: a reader must always be able to enlarge the page. Two ways to
    break it, both of which have bitten this file:

    1. `touch-action: pan-y` or `none`. These read as "let the browser scroll but not
       pan sideways" and quietly withhold two-finger zoom as well. The molecule sheet
       covers ~85% of a phone screen, so once it did that, a reader who had zoomed in
       had almost nowhere to place two fingers to zoom back out. Every gesture this
       file drives is single-finger, so naming pinch-zoom costs nothing.
    2. `user-scalable=no` / `maximum-scale=1` in the viewport meta. Never add these.
    """
    for meta in re.findall(r'<meta name="viewport"[^>]*>', src):
        for banned in ('user-scalable=no', 'user-scalable=0', 'maximum-scale=1'):
            if banned in meta.replace(' ', ''):
                fail('pinch', 'viewport meta forbids zoom (%s) — readers cannot enlarge '
                              'the page' % banned)

    ok_alone = {'auto', 'manipulation'}   # both already permit pinch-zoom
    offenders = []
    for m in re.finditer(r'([^\n{}]+)\{([^}]*?)touch-action\s*:\s*([^;}]+)', src):
        sel, val = m.group(1).strip(), m.group(3).strip()
        if val in ok_alone or 'pinch-zoom' in val:
            continue
        offenders.append('%s {touch-action:%s}' % (sel.split('\n')[-1].strip()[:52], val))
    if offenders:
        fail('pinch', 'these rules withhold pinch-zoom — add "pinch-zoom" to the value: %s'
                      % '; '.join(offenders[:3]))

    n = len(re.findall(r'touch-action\s*:', strip_comments(src)))
    note('pinch-zoom: viewport allows scaling; all %d touch-action rules permit two-finger '
         'zoom' % n)


def check_unwrapped_counts(src):
    """Counts describing V2's own data must be DERIVED, never typed into prose.

    check_v2fact_keys validates every <span data-v2fact> and its fallback. The gap it cannot
    see is a number typed straight into a sentence and never wrapped — "16 molecules engage
    CB1" sailed through for months while the data said 17. This closes that gap.

    Anything inside a data-v2fact element is derived (or is its own checked fallback), so
    those elements are removed before scanning. General statements about the plant rather
    than about V2's tables ("over 100 cannabinoids identified in cannabis") are not counts
    of anything here, so hedged numbers are skipped.
    """
    body = strip_comments(src)
    body = re.sub(r'<script.*?</script>', '', body, flags=re.S)
    # drop derived elements wholesale — span for scalars, div for the group paragraphs
    body = re.sub(r'<span data-v2fact=.*?</span>', ' ', body, flags=re.S)
    body = re.sub(r'<div data-v2fact=.*?</div>', ' ', body, flags=re.S)

    nouns = r'(?:health\s+)?(?:conditions?|topics?|molecules?|cannabinoids?|terpenes?|flavonoids?)'
    hedged = re.compile(r'(?:over|about|approximately|more than|nearly|some|than)\s*$', re.I)
    bad = []
    for m in re.finditer(r'(\d{2,3})\s*' + nouns, body, re.I):
        before = body[max(0, m.start() - 30):m.start()]
        if hedged.search(before):
            continue                       # a fact about the plant, not a count of our data
        if re.search(r'from\s+\d{2,3}\s+to\s*$', before):
            continue                       # "expanded from 30 to <derived>"
        bad.append(re.sub(r'\s+', ' ', body[max(0, m.start() - 45): m.end() + 5]).strip())
    if bad:
        fail('derived counts',
             'these counts are typed into prose instead of derived from the data '
             '(wrap in <span data-v2fact="...">): %s' % ' | '.join(bad[:3]))
    else:
        note('derived counts: every count describing the data is derived, none hardcoded')


def check_pubmed_links(src, data):
    """UX-77: no PubMed link may open a search that returns nothing.

    V2 is offline-first and cannot ask NCBI at render time, so every search the file can
    build was probed against the E-utilities API and the empty ones recorded in
    PUBMED_EMPTY. Links for those are replaced by a plain statement of absence.

    This check cannot re-run the probe (preflight is offline by default). What it can do is
    notice when the probe has gone out of date: if a molecule, indication, condition or
    molecule-condition pairing has been added since, its searches have never been verified
    and may be dead on arrival. The manifest makes that detectable instead of silent.
    """
    if 'const PUBMED_EMPTY' not in src:
        fail('pubmed links', 'PUBMED_EMPTY is missing — dead-link suppression is not wired')
        return
    blk = src[src.index('const PUBMED_EMPTY'):]
    blk = blk[:blk.index('\n};')]
    m = re.search(r"probed:\s*\{([^}]*)\}", blk)
    if not m:
        fail('pubmed links', 'PUBMED_EMPTY has no coverage manifest')
        return
    claimed = {k: int(v) for k, v in re.findall(r"(\w+):\s*(\d+)", m.group(1))}

    mols = data['M']
    conds = [c for c in data['CONDITIONS'] if c.get('id') != 'cancer-adverse']
    n_pairs = sum(len(mols[k].get('indications') or []) for k in mols)
    n_combo = sum(len(c.get('molecules') or {}) for c in data['CONDITIONS'])
    actual = {'pairs': n_pairs, 'mols': len(mols),
              'conds': len(conds) + 1, 'combos': n_combo}   # +1: the adverse filter renders too

    for key in ('pairs', 'mols', 'conds', 'combos'):
        if claimed.get(key) != actual[key]:
            fail('pubmed links',
                 'content changed since the PubMed probe (%s: probed %s, now %s) — re-run the '
                 'probe and refresh PUBMED_EMPTY, or new links may open empty searches'
                 % (key, claimed.get(key), actual[key]))

    for kind in ('pair', 'mol', 'ent', 'syn', 'combo'):
        if not re.search(kind + r':\s*\[', blk):
            fail('pubmed links', 'PUBMED_EMPTY.%s is missing' % kind)
    if '[ti]' in strip_comments(src):
        fail('pubmed links', 'a title-only [ti] search is back; use [tiab] (see UX-76)')

    n = sum(len(re.findall(r"'", re.search(k + r':\s*\[(.*?)\]', blk, re.S).group(1))) // 2
            for k in ('pair', 'mol', 'ent', 'syn', 'combo')
            if re.search(k + r':\s*\[(.*?)\]', blk, re.S))
    note('pubmed links: %d verified-empty searches suppressed; probe covers %d pairs, '
         '%d molecules, %d conditions, %d combinations'
         % (n, claimed['pairs'], claimed['mols'], claimed['conds'], claimed['combos']))


def check_ddi_reachable(data):
    """DRUG-09: every drug-drug pair must be able to fire, and no drug id may be duplicated.

    Reads the EVALUATED tables, not the source text. The first version of this check pattern-matched
    index.html and reported 11 unreachable pairs that were all fine — DI_DATA is written in two
    styles, single-quoted JS and double-quoted JSON, and the pattern only saw the first. The same
    mistake, in the same shape, as the citation regex that once produced a false "42 of 59
    conditions uncited". If a check can be written against the real objects, write it that way.

    Two real defects this catches:
      * a pair naming a drug with no DI_DATA record can never fire — the medication list is built
        from DI_DATA ids, so an absent id can never appear in medIds;
      * a duplicated drug id makes idToIndex resolve to whichever entry came last, so one of the
        two records becomes unreachable and the interaction shown may not be the one curated for
        that drug. Caught exactly once, on a duplicate this session's own author introduced.
    """
    di, ddi = data.get('DI_DATA') or [], data.get('DDI_DATA') or []
    if not di or not ddi:
        fail('ddi reachable', 'DI_DATA / DDI_DATA did not evaluate')
        return
    seen, dupes = {}, []
    for i, e in enumerate(di):
        rid = e.get('id')
        if rid in seen:
            dupes.append('%s (entries %d and %d)' % (rid, seen[rid], i))
        else:
            seen[rid] = i
    for d in dupes:
        fail('ddi reachable', 'duplicate DI_DATA drug id: %s — one record is unreachable' % d)
    dead = []
    for pair in ddi:
        groups = pair.get('groups')
        if groups:
            ok = all(any(x in seen for x in g) for g in groups if g)
        else:
            drugs = pair.get('drugs') or []
            # without explicit groups the matcher tests only the first two ids
            ok = len(drugs) >= 2 and drugs[0] in seen and drugs[1] in seen
        if not ok:
            missing = [x for x in (pair.get('drugs') or []) if x not in seen]
            dead.append('%s (missing %s)' % (pair.get('id'), ', '.join(missing) or 'a required drug'))
    for d in dead:
        fail('ddi reachable', 'drug-drug pair can never fire: %s' % d)
    # a pair with >2 drugs and no groups silently ignores everything past the second
    flat = [p.get('id') for p in ddi
            if not p.get('groups') and len(p.get('drugs') or []) > 2]
    for f in flat:
        fail('ddi reachable',
             'pair %s lists more than two drugs without `groups` — elements 3+ are ignored by the '
             'matcher' % f)
    if not dupes and not dead and not flat:
        note('ddi reachable: %d drugs, %d pairs, all pairs can fire, no duplicate ids'
             % (len(di), len(ddi)))


def check_backlog():
    """The backlog is the plan — duplicate IDs make a row unfindable. Two pairs shipped."""
    path = os.path.join(HERE, 'Project_Backlog.xlsx')
    if not os.path.exists(path):
        note('backlog: Project_Backlog.xlsx not present, skipped')
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        note('backlog: openpyxl unavailable, skipped')
        return
    ws = load_workbook(path)['Backlog']
    hdr = [c.value for c in ws[1]]
    ci = {h: n + 1 for n, h in enumerate(hdr) if h}
    ids, bad_status = [], []
    valid = {'Not Started', 'In Progress', 'Blocked', 'Done', 'Deferred'}
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(r, ci['ID']).value
        if not rid:
            continue
        ids.append(str(rid))
        st = ws.cell(r, ci['Status']).value
        if st not in valid:
            bad_status.append('%s -> %r' % (rid, st))
    dupes = sorted({i for i in ids if ids.count(i) > 1})
    if dupes:
        fail('backlog', 'duplicate row IDs: %s' % dupes)
    if bad_status:
        fail('backlog', 'invalid Status value(s): %s' % bad_status[:8])
    note('backlog: %d rows, ids unique, statuses valid' % len(ids))


ALLOWED_SOURCES = {'label'}   # DRUG-26: FDA-approved prescribing information

# DRUG-26: the attribution rule below is enforced for the WHOLE database. Thirteen entries were in
# violation when the guard was written (QA Pass A, 2026-08-30). DRUG-27 closed the twelve Group 3
# entries — nine by citing the literature, three by attributing to FDA labeling. The one that
# remains is a GRADE question, not a citation question, and is held for owner review.
#
# This register is self-cleaning: check_evidence_attribution FAILS if an id listed here no longer
# violates, so it cannot rot into a stale allowlist the way index.html's line numbers did. Removing
# the last entry removes the register.
ATTRIBUTION_OPEN = {
    # Group 4 — the GRADE itself is the open question, not the citation. Pending owner review.
    'hydroxyzine_cns': 'pharmacodynamic inference graded B; B->C recommended, or attribute to label',
}


def check_record_schema(data):
    """DRUG-24: a drug record and a drug-drug pair must each live in their OWN array.

    This guard exists because the check it replaces did not exist, and the gap shipped a real
    defect during this very session. Seven DDI-shaped objects were appended to DI_DATA instead of
    DDI_DATA — the insertion script bracket-searched for the literal "\n  ];" after the last pair,
    but DDI_DATA closes on "\n];" with NO indentation while DI_DATA closes with two spaces, so the
    search ran straight past the end of DDI_DATA and landed on DI_DATA's terminator instead.

    Every existing check passed. check_ddi_reachable counted "284 drugs, 96 pairs" and called it
    clean, because it only asks whether ids resolve — never whether an entry is the right SHAPE.
    Those seven objects would have rendered in the medication list as drugs with no name, no brands,
    no class and no molecules. The count was the only visible symptom, and a count is easy to skim.

    So: assert the shape, not just the references. A DI record must carry the keys the card renderer
    reads; a DDI pair must carry its own. Neither may carry the other's distinctive keys.
    """
    di, ddi = data.get('DI_DATA') or [], data.get('DDI_DATA') or []
    di_required = ('id', 'drug', 'brands', 'cls', 'cat', 'mols', 'sev', 'ev',
                   'mech', 'effect', 'monitor', 'pmid')
    ddi_required = ('id', 'drugs', 'sev', 'ev', 'label', 'mech', 'effect', 'monitor', 'pmid')
    di_only = ('drug', 'brands', 'cls', 'cat', 'mols')
    ddi_only = ('drugs', 'groups', 'label')
    for e in di:
        rid = e.get('id', '?')
        missing = [k for k in di_required if k not in e]
        if missing:
            fail('record schema',
                 'DI_DATA record "%s" is missing %s — is it a drug-drug pair in the wrong array?'
                 % (rid, ', '.join(missing)))
        stray = [k for k in ddi_only if k in e]
        if stray:
            fail('record schema',
                 'DI_DATA record "%s" carries pair-only key(s) %s — it belongs in DDI_DATA'
                 % (rid, ', '.join(stray)))
    for e in ddi:
        rid = e.get('id', '?')
        missing = [k for k in ddi_required if k not in e]
        if missing:
            fail('record schema',
                 'DDI_DATA pair "%s" is missing %s — is it a drug record in the wrong array?'
                 % (rid, ', '.join(missing)))
        stray = [k for k in di_only if k in e]
        if stray:
            fail('record schema',
                 'DDI_DATA pair "%s" carries drug-only key(s) %s — it belongs in DI_DATA'
                 % (rid, ', '.join(stray)))
    # severity and evidence are closed vocabularies; a typo silently mis-sorts and mis-badges
    for label, rows in (('DI_DATA', di), ('DDI_DATA', ddi)):
        for e in rows:
            if e.get('sev') not in ('major', 'moderate', 'minor'):
                fail('record schema', '%s "%s" has sev=%r' % (label, e.get('id'), e.get('sev')))
            if e.get('ev') not in ('A', 'B', 'C', 'D'):
                fail('record schema', '%s "%s" has ev=%r' % (label, e.get('id'), e.get('ev')))
    # DRUG-26: `source` names the AUTHORITY behind a statement when that authority is a
    # regulatory document rather than a paper. It is orthogonal to `ev`, which keeps describing
    # the strength and type of evidence — a label contraindication rests on human data the
    # manufacturer submitted, so ev:'A' with source:'label' is coherent, not a contradiction.
    for label, rows in (('DI_DATA', di), ('DDI_DATA', ddi)):
        for e in rows:
            if 'source' in e and e['source'] not in ALLOWED_SOURCES:
                fail('record schema',
                     '%s "%s" has source=%r — allowed: %s'
                     % (label, e.get('id'), e.get('source'), sorted(ALLOWED_SOURCES)))
    note('record schema: %d drug records and %d pairs each carry the right keys for their array'
         % (len(di), len(ddi)))


def n_ab(data):
    return sum(1 for rows in (data.get('DI_DATA') or [], data.get('DDI_DATA') or [])
               for e in rows if e.get('ev') in ('A', 'B'))


def check_evidence_attribution(data):
    """DRUG-26: an A- or B-graded entry ASSERTS human evidence, so it must say where that comes from.

    QA Pass A (2026-08-30) found 36 entries graded A or B carrying no citation at all — 30 of the 90
    A/B pairs. Nothing caught it: check_pmids verifies that PMIDs which are PRESENT resolve, and never
    asks whether a claim of human evidence has any attribution behind it. The entries rendered an
    empty citation slot, which reads as "unsupported" for exactly the records with the strongest
    regulatory backing.

    The rule is deliberately narrow. Grades C (preclinical) and D (theoretical) are exempt: a record
    that says "no human data exists" is honest without a citation, and 160 of them legitimately have
    none. Only A and B — the grades that claim humans were studied — must point somewhere, either at
    a PMID or at a named authority such as the FDA label.
    """
    offenders = []
    for label, rows in (('DI_DATA', data.get('DI_DATA') or []),
                        ('DDI_DATA', data.get('DDI_DATA') or [])):
        for e in rows:
            if e.get('ev') not in ('A', 'B'):
                continue
            has_pmid = bool(re.search(r'\d{5,8}', str(e.get('pmid') or '')))
            if has_pmid or e.get('source') in ALLOWED_SOURCES:
                continue
            offenders.append((e.get('id'), '%s "%s" (%s/%s)'
                             % (label, e.get('id'), e.get('sev'), e.get('ev'))))
    new_offenders = [o for o in offenders if o[0] not in ATTRIBUTION_OPEN]
    for _, desc in new_offenders:
        fail('evidence attribution',
             '%s is graded A/B — which asserts human evidence — but cites no PMID and names no '
             'source' % desc)
    # a register entry that no longer violates must be REMOVED, or the register goes stale
    still = {o[0] for o in offenders}
    fixed = sorted(set(ATTRIBUTION_OPEN) - still)
    for rid in fixed:
        fail('evidence attribution',
             '"%s" is listed in ATTRIBUTION_OPEN but now has attribution — delete its line' % rid)
    if not new_offenders and not fixed:
        note('evidence attribution: %d A/B entries all attributed, except %d registered as open '
             '(Groups 3-4)' % (n_ab(data), len(still)))
    if not offenders:
        note('evidence attribution: all %d A/B-graded entries cite a PMID or name a source'
             % n_ab(data))


def check_hasrisk_invariant(data):
    """DRUG-24: hasRisk is editorial intent; ADVERSE_FINDINGS is what the reader actually sees.

    hasRisk is written on conditions but read by NO code — the caution a user sees comes from
    condHasAdverse(), which matches an adverse finding by condition label or by group. That makes
    hasRisk a comment that looks like data, and it had already drifted once: an earlier note in
    index.html claimed hasRisk was what flagged a condition, which was wrong.

    Deleting it would throw away a real editorial signal, so it is enforced instead. The rule is
    one-directional and deliberately so: if a curator marks a condition risky, the reader MUST see
    a caution on it. The converse is not required — condHasAdverse legitimately matches by group
    and covers more conditions than hasRisk marks.
    """
    conds = [c for c in data['CONDITIONS'] if c.get('id') != 'cancer-adverse']
    adverse = data.get('ADVERSE_FINDINGS') or {}

    def shown(cond):
        if cond.get('isRisk'):
            return True                       # the global filter row is the adverse entry point
        groups = group_keys(cond.get('group'))
        for mol in adverse:
            for f in adverse[mol]:
                if f.get('footnote'):
                    continue
                if f.get('condition') == cond.get('label') or f.get('group') in groups:
                    return True
        return False

    flagged = [c for c in conds if c.get('hasRisk')]
    silent = [c.get('id') for c in flagged if not shown(c)]
    for cid in silent:
        fail('hasRisk invariant',
             'condition "%s" is marked hasRisk:true but no adverse finding matches it by label or '
             'group — the curator flagged a risk the reader never sees' % cid)
    if not silent:
        note('hasRisk invariant: all %d flagged conditions surface a caution to the reader'
             % len(flagged))


# ── driver ─────────────────────────────────────────────────────────────────────
def main():
    online = '--online' in sys.argv
    src = open(SRC, encoding='utf-8').read()
    print('preflight — index.html (%d KB)%s\n' % (len(src) // 1024,
                                                 '  [--online]' if online else ''))
    data = load_data(src)

    check_scripts_parse(src)
    check_v2fact_keys(src, data)
    check_demo_guided_parity(src, data)
    check_entities(src)
    check_banned_terms(src)
    check_brand_lockups(src)
    check_standing_rules(src)
    check_data_integrity(src, data)
    check_faq_parity(src)
    check_terminology(src)
    check_recommendation_wording(src)
    check_greek_notation(src)
    check_pinch_zoom(src)
    check_unwrapped_counts(src)
    check_pubmed_links(src, data)
    check_pmids(src, online)
    check_ddi_reachable(data)
    check_record_schema(data)
    check_hasrisk_invariant(data)
    check_evidence_attribution(data)
    check_backlog()

    for n in NOTES:
        print('  ·', n)
    if FAILURES:
        print('\n%d PROBLEM(S) — do not deploy:\n' % len(FAILURES))
        width = max(len(c) for c, _ in FAILURES)
        for check, msg in FAILURES:
            print('  ✗ [%s] %s' % (check.ljust(width), msg))
        return 1
    print('\nAll preflight checks passed — safe to deploy.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
